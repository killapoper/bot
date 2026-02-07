import os
import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from datetime import datetime

class ExcelManager:
    def __init__(self, filename):
        self.filename = filename
        self.ensure_file_exists()

    def ensure_file_exists(self):
        """Creates the Excel file with headers if it doesn't exist."""
        if not os.path.exists(self.filename):
            wb = Workbook()
            ws = wb.active
            ws.title = "Purchases"
            # Headers based on requirements
            headers = [
                "ID Записи", "Дата создания", "User ID", "Имя/Должность", 
                "Тип закупки", "Дата закупки (Лейбл)", "Дата закупки (Значение)",
                "Организация", "Контакт", "Отрасль", "Получатель",
                "Позиция: Название", "Позиция: Цена", 
                "Позиция: Фото (Превью)", "Позиция: Фото (Ссылка)", "Чек", "1C"
            ]
            ws.append(headers)
            # Set column widths
            ws.column_dimensions['N'].width = 15 # Photo Preview
            ws.column_dimensions['O'].width = 15 # Photo Link
            ws.column_dimensions['P'].width = 15 # Receipt
            ws.column_dimensions['Q'].width = 10 # 1C
            wb.save(self.filename)
            wb.close()

    def add_purchase(self, user_id, purchase_data):
        """
        Adds purchase data to the Excel file.
        """
        try:
            wb = openpyxl.load_workbook(self.filename)
            ws = wb.active
            
            created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            purchase_id = ws.max_row  # Simple ID strategy

            # Receipt Hyperlink Logic
            receipt_url = purchase_data.get('receipt_url')
            receipt_path = purchase_data.get('receipt_path')
            
            receipt_cell_value = "Нет"
            
            if receipt_url:
                # Use Public Google Drive Link
                receipt_cell_value = f'=HYPERLINK("{receipt_url}", "Просмотр")'
            elif receipt_path:
                # Fallback to local path (absolute)
                abs_path = os.path.abspath(receipt_path)
                receipt_cell_value = f'=HYPERLINK("{abs_path}", "Локальный файл")'

            for pos in purchase_data.get('positions', []):
                # Photo Column Hyperlink Logic (Column O)
                photo_url = pos.get('photo_url')
                photo_link_value = ""
                if photo_url:
                    photo_link_value = f'=HYPERLINK("{photo_url}", "Просмотр")'

                # Basic data row
                row_data = [
                    purchase_id,
                    created_at,
                    user_id,
                    purchase_data.get('user_name', 'Не указано'),
                    purchase_data.get('type'),
                    purchase_data.get('date_label'),
                    purchase_data.get('date_value'),
                    purchase_data.get('organization'),
                    purchase_data.get('contact'),
                    purchase_data.get('industry'),
                    purchase_data.get('recipient'),
                    pos.get('name'),
                    pos.get('price'),
                    "",                 # N: Photo Preview Placeholder
                    photo_link_value,   # O: Photo Link
                    receipt_cell_value, # P: Receipt
                    purchase_data.get('1c_status', 'Нет') # Q: 1C Status
                ]
                ws.append(row_data)
                
                # Insert Image Preview (Column N)
                photo_path = pos.get('photo_path')
                if photo_path and os.path.exists(photo_path):
                    try:
                        img = Image(photo_path)
                        img.width = 100
                        img.height = 100
                        current_row = ws.max_row
                        ws.add_image(img, f"N{current_row}")
                        ws.row_dimensions[current_row].height = 80
                    except Exception as img_err:
                        print(f"Error adding image: {img_err}")

            self.save_workbook(wb)
            return purchase_id
            
        except Exception as e:
            print(f"Error saving to Excel: {e}")
            return None

    def save_workbook(self, wb):
        """Saves the workbook with retry logic."""
        import time
        retries = 3
        for i in range(retries):
            try:
                wb.save(self.filename)
                wb.close()
                return True
            except PermissionError:
                if i < retries - 1:
                    print(f"File locked. Retrying in 1s... ({i+1}/{retries})")
                    time.sleep(1)
                else:
                    print("Error: Could not save Excel file. It might be open.")
                    raise
            except Exception as e:
                 print(f"Save error: {e}")
                 wb.close()
                 raise
            


    def clear_data(self):
        """Clears all data by overwriting with a fresh workbook."""
        import time
        retries = 3
        for i in range(retries):
            try:
                # Instead of removing the file (which often fails if any handle is open),
                # we just create a new workbook and save it OVER the existing one.
                wb = Workbook()
                ws = wb.active
                ws.title = "Purchases"
                headers = [
                    "ID Записи", "Дата создания", "User ID", "Имя/Должность", 
                    "Тип закупки", "Дата закупки (Лейбл)", "Дата закупки (Значение)",
                    "Организация", "Контакт", "Отрасль", "Получатель",
                    "Позиция: Название", "Позиция: Цена", 
                    "Позиция: Фото (Превью)", "Позиция: Фото (Ссылка)", "Чек", "1C"
                ]
                ws.append(headers)
                # Set column widths
                ws.column_dimensions['N'].width = 15 # Photo Preview
                ws.column_dimensions['O'].width = 15 # Photo Link
                ws.column_dimensions['P'].width = 15 # Receipt
                ws.column_dimensions['Q'].width = 10 # 1C
                
                wb.save(self.filename)
                wb.close()
                return True
            except PermissionError:
                if i < retries - 1:
                    print(f"File locked (clear). Retrying in 1s... ({i+1}/{retries})")
                    time.sleep(1)
                else:
                    print("Error: Could not clear Excel file. It might be open.")
            except Exception as e:
                print(f"Error clearing Excel file: {e}")
                return False
        return False

    def get_last_purchases(self, limit=5):
        """Returns the last `limit` purchases from the Excel file."""
        wb = None
        try:
            wb = openpyxl.load_workbook(self.filename)
            ws = wb.active
            
            # Get data rows (skip header)
            # Use values_only=True so we don't hold references to cells
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            
            if not rows:
                return []
            
            # Sort by ID (first column) descending or just take last rows if modification date isn't reliable
            # Assuming append adds to bottom, last rows are newest.
            last_rows = rows[-limit:]
            last_rows.reverse() # Show newest first
            
            purchases = []
            for row in last_rows:
                # Row structure based on add_purchase:
                # 0: ID, 1: CreatedAt, 2: UserID, 3: Name, 4: Type, 5: DateLabel, 6: DateValue
                # 7: Org, 8: Contact, 9: Industry, 10: Recipient, 11: PosName, 12: PosPrice, 13: Photo
                
                # Careful with index out of range if row is short
                def get_col(idx):
                    return row[idx] if idx < len(row) else None

                purchases.append({
                    'id': get_col(0),
                    'created_at': get_col(1),
                    'user': get_col(3),
                    'org': get_col(7),
                    'position': get_col(11),
                    'price': get_col(12)
                })
                
            return purchases
            
        except Exception as e:
            print(f"Error reading Excel file: {e}")
            return []
        finally:
            if wb:
                wb.close()
